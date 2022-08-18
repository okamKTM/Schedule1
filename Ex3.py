import os
import openpyxl
import glob

wb = openpyxl.Workbook()
wb.create_sheet()
Sheet = wb.sheetnames
print(Sheet)

sheet1 = wb['Sheet1']
Sheet = wb.remove(sheet1)
print(Sheet)

wb.save('test.xlsx')
print(glob.glob("*.xlsx")) #拡張子が.xlsxのものを探索


